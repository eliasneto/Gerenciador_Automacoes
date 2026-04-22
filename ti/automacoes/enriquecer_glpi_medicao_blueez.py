import json
import os
import shutil
import re
import time
from datetime import datetime
from pathlib import Path

import MySQLdb
import pandas as pd
from openpyxl import load_workbook


VERSAO_AUTOMACAO = "1.0.0"
COLUNA_NUMERO_BLUEEZ = "numero_solicitacao"
COLUNA_CHAMADOS_GLPI = "Chamados GLPI"
COLUNA_CHAMDOS_GLPI_ANTIGA = "Chamdos GLPI"
COLUNA_SOLICITANTE_GLPI = "Solicitante GLPI"
COLUNA_DATA_CRIACAO_GLPI = "Data Criacao GLPI"
GLPI_DB_CONFIG_PADRAO = {
    "host": "192.168.90.250",
    "port": 3306,
    "user": "ti",
    "password": "Q1w2e3r4$",
    "database": "orquestrador",
    "table": "blueez_medicoes_monitoradas",
}
MYSQL_ERROS_CONEXAO = {2002, 2003, 2005, 2006, 2013}


class AbortarAutomacao(Exception):
    pass


def executar(
    input_path=None,
    input_paths=None,
    attachments=None,
    output_dir=None,
    should_stop=None,
    log=None,
    parametros="",
    parametros_json=None,
):
    logger = log or (lambda message: None)
    logger(f"Enriquecer GLPI Medicao BlueEZ - Versao {VERSAO_AUTOMACAO}")

    arquivos_principais = [Path(path) for path in (input_paths or ([] if input_path is None else [input_path]))]
    pasta_saida = Path(output_dir or Path.cwd())
    pasta_saida.mkdir(parents=True, exist_ok=True)

    if should_stop:
        should_stop()

    parametros_execucao = carregar_parametros(parametros, parametros_json)
    config_db = carregar_config_glpi(parametros_execucao)
    planilha_entrada = selecionar_planilha(arquivos_principais, input_path)
    planilha_saida = copiar_planilha(planilha_entrada, pasta_saida)
    logger(f"Planilha de trabalho criada: {planilha_saida.name}")

    df = pd.read_excel(planilha_saida)
    validar_planilha(df)

    numeros_blueez = carregar_numeros_blueez_pendentes(df)
    logger(f"Foram encontrados {len(numeros_blueez)} numero(s) BlueEZ pendente(s) para consulta.")
    mapa_glpi, aviso_consulta = consultar_dados_glpi(numeros_blueez, config_db, logger)
    atualizar_planilha_saida(planilha_saida, mapa_glpi, logger)
    logger("Planilha de saida enriquecida com os dados do GLPI.")

    mensagem = (
        f"Enriquecimento GLPI concluido. Arquivo gerado: {planilha_saida.name}. "
        f"Registros consultados: {len(numeros_blueez)}. Registros encontrados no banco: {len(mapa_glpi)}."
    )
    if aviso_consulta:
        mensagem += f" Aviso: {aviso_consulta}"

    return {
        "message": mensagem
    }


def carregar_parametros(parametros, parametros_json):
    if isinstance(parametros_json, dict):
        return dict(parametros_json)
    texto = (parametros or "").strip()
    if not texto:
        return {}
    try:
        valor = json.loads(texto)
        return valor if isinstance(valor, dict) else {"raw": valor}
    except json.JSONDecodeError:
        return {"raw": texto}


def carregar_config_glpi(parametros_execucao):
    return {
        "host": (
            parametros_execucao.get("local_db_host")
            or os.getenv("LOCAL_DB_HOST")
            or GLPI_DB_CONFIG_PADRAO["host"]
        ),
        "port": int(
            parametros_execucao.get("local_db_port")
            or os.getenv("LOCAL_DB_PORT")
            or GLPI_DB_CONFIG_PADRAO["port"]
        ),
        "user": (
            parametros_execucao.get("local_db_user")
            or os.getenv("LOCAL_DB_USER")
            or GLPI_DB_CONFIG_PADRAO["user"]
        ),
        "password": (
            parametros_execucao.get("local_db_pass")
            or os.getenv("LOCAL_DB_PASS")
            or GLPI_DB_CONFIG_PADRAO["password"]
        ),
        "database": (
            parametros_execucao.get("local_db_name")
            or os.getenv("LOCAL_DB_NAME")
            or GLPI_DB_CONFIG_PADRAO["database"]
        ),
        "table": (
            parametros_execucao.get("local_db_table")
            or os.getenv("LOCAL_DB_TABLE")
            or GLPI_DB_CONFIG_PADRAO["table"]
        ),
    }


def selecionar_planilha(arquivos_principais, input_path):
    candidatos = list(arquivos_principais)
    if input_path:
        candidatos.insert(0, Path(input_path))
    for arquivo in candidatos:
        if arquivo.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
            return arquivo
    raise AbortarAutomacao("Nenhuma planilha Excel foi enviada como arquivo principal.")


def copiar_planilha(planilha_entrada, pasta_saida):
    destino = pasta_saida / f"glpi_enriquecido_{datetime.now().strftime('%Y%m%d_%H%M%S')}{planilha_entrada.suffix}"
    shutil.copy2(planilha_entrada, destino)
    return destino


def validar_planilha(df):
    colunas = {str(coluna).strip().lower() for coluna in df.columns}
    if COLUNA_NUMERO_BLUEEZ not in colunas:
        raise AbortarAutomacao(
            "A planilha informada nao possui a coluna 'numero_solicitacao'. "
            "Use como entrada a planilha de saida da automacao Medicao BlueEZ."
        )


def carregar_numeros_blueez_pendentes(df):
    coluna_numero = None
    coluna_glpi = None

    for coluna in df.columns:
        nome_coluna = str(coluna).strip().lower()
        if nome_coluna == COLUNA_NUMERO_BLUEEZ:
            coluna_numero = coluna
        elif nome_coluna in {
            COLUNA_CHAMADOS_GLPI.lower(),
            COLUNA_CHAMDOS_GLPI_ANTIGA.lower(),
        }:
            coluna_glpi = coluna

    if coluna_numero is None:
        return []

    numeros = []
    for _, linha in df.iterrows():
        numero = normalizar_numero_blueez(linha.get(coluna_numero))
        glpi_atual = linha.get(coluna_glpi) if coluna_glpi is not None else ""
        if numero and not possui_valor(glpi_atual):
            numeros.append(numero)
    return sorted(set(numeros))


def consultar_dados_glpi(numeros_blueez, config_db, logger):
    if not numeros_blueez:
        logger("Nenhum numero BlueEZ valido foi encontrado para consulta.")
        return {}, ""

    placeholders = ", ".join(["%s"] * len(numeros_blueez))
    tabela = config_db["table"]
    query = f"""
        SELECT
            CAST(medicao_id AS CHAR) AS medicao_id,
            CAST(glpi_ticket_id AS CHAR) AS glpi_ticket_id,
            solicitante_login,
            created_at
        FROM {tabela}
        WHERE REPLACE(REPLACE(CAST(medicao_id AS CHAR), '.', ''), ',', '') IN ({placeholders})
        ORDER BY created_at DESC
    """

    total_tentativas = 3
    ultimo_erro = None
    linhas = []
    for tentativa in range(1, total_tentativas + 1):
        conexao = None
        cursor = None
        try:
            logger(
                "Consultando base GLPI em "
                f"{config_db['host']}:{config_db['port']} / {config_db['database']}.{config_db['table']} "
                f"(tentativa {tentativa}/{total_tentativas})."
            )
            conexao = MySQLdb.connect(
                host=config_db["host"],
                port=int(config_db["port"]),
                user=config_db["user"],
                passwd=config_db["password"],
                db=config_db["database"],
                charset="utf8mb4",
                use_unicode=True,
                connect_timeout=10,
                read_timeout=20,
                write_timeout=20,
            )
            cursor = conexao.cursor()
            cursor.execute(query, numeros_blueez)
            linhas = cursor.fetchall()
            ultimo_erro = None
            break
        except MySQLdb.OperationalError as exc:
            ultimo_erro = exc
            codigo_erro = exc.args[0] if exc.args else None
            logger(f"Falha de conexao com a base GLPI na tentativa {tentativa}: {exc}")
            if codigo_erro not in MYSQL_ERROS_CONEXAO or tentativa >= total_tentativas:
                break
            time.sleep(3)
        except Exception as exc:
            ultimo_erro = exc
            logger(f"Falha inesperada ao consultar a base GLPI na tentativa {tentativa}: {exc}")
            break
        finally:
            try:
                if cursor is not None:
                    cursor.close()
            except Exception:
                pass
            try:
                if conexao is not None:
                    conexao.close()
            except Exception:
                pass

    if ultimo_erro is not None:
        aviso = (
            "Nao foi possivel consultar a base GLPI nesta execucao. "
            "A planilha foi gerada com as colunas GLPI em branco para nova tentativa posterior."
        )
        logger(aviso)
        return {}, aviso

    mapa = {}
    for medicao_id, glpi_ticket_id, solicitante_login, created_at in linhas:
        chave = normalizar_numero_blueez(medicao_id)
        if not chave or chave in mapa:
            continue
        mapa[chave] = {
            "glpi_ticket_id": "" if glpi_ticket_id is None else str(glpi_ticket_id).strip(),
            "solicitante_login": "" if solicitante_login is None else str(solicitante_login).strip(),
            "created_at": formatar_data_glpi(created_at),
        }

    logger(f"Consulta concluida. {len(mapa)} registro(s) encontrado(s) no banco.")
    return mapa, ""


def atualizar_planilha_saida(caminho_xlsx, mapa_glpi, logger):
    wb = load_workbook(caminho_xlsx)
    ws = wb.active

    col_numero = garantir_coluna_existente(ws, COLUNA_NUMERO_BLUEEZ)
    col_glpi = garantir_coluna_chamados_glpi(ws)
    col_solicitante = garantir_coluna(ws, COLUNA_SOLICITANTE_GLPI)
    col_data_criacao = garantir_coluna(ws, COLUNA_DATA_CRIACAO_GLPI)

    encontrados = 0
    pulados = 0
    for linha in range(2, ws.max_row + 1):
        numero = normalizar_numero_blueez(ws.cell(row=linha, column=col_numero).value)
        if not numero:
            continue
        glpi_existente = ws.cell(row=linha, column=col_glpi).value
        if possui_valor(glpi_existente):
            pulados += 1
            continue
        dados = mapa_glpi.get(numero, {})
        ws.cell(row=linha, column=col_glpi).value = dados.get("glpi_ticket_id", "")
        ws.cell(row=linha, column=col_solicitante).value = dados.get("solicitante_login", "")
        ws.cell(row=linha, column=col_data_criacao).value = dados.get("created_at", "")
        if dados:
            encontrados += 1

    wb.save(caminho_xlsx)
    logger(
        "Planilha atualizada. "
        f"{encontrados} linha(s) receberam dados do GLPI e "
        f"{pulados} linha(s) foram mantidas porque ja tinham ID GLPI."
    )


def garantir_coluna_existente(ws, nome):
    for coluna in range(1, ws.max_column + 1):
        valor = ws.cell(row=1, column=coluna).value
        if valor is not None and str(valor).strip().lower() == nome.lower():
            return coluna
    raise AbortarAutomacao(f"A coluna obrigatoria '{nome}' nao foi encontrada na planilha.")


def garantir_coluna(ws, nome):
    for coluna in range(1, ws.max_column + 1):
        valor = ws.cell(row=1, column=coluna).value
        if valor is not None and str(valor).strip().lower() == nome.lower():
            return coluna
    coluna = ws.max_column + 1
    ws.cell(row=1, column=coluna).value = nome
    return coluna


def garantir_coluna_chamados_glpi(ws):
    for coluna in range(1, ws.max_column + 1):
        valor = ws.cell(row=1, column=coluna).value
        if valor is None:
            continue
        nome = str(valor).strip().lower()
        if nome == COLUNA_CHAMADOS_GLPI.lower():
            return coluna
        if nome == COLUNA_CHAMDOS_GLPI_ANTIGA.lower():
            ws.cell(row=1, column=coluna).value = COLUNA_CHAMADOS_GLPI
            return coluna
    coluna = ws.max_column + 1
    ws.cell(row=1, column=coluna).value = COLUNA_CHAMADOS_GLPI
    return coluna


def normalizar_numero_blueez(valor):
    return re.sub(r"\D+", "", str(valor or "")).strip()


def possui_valor(valor):
    if valor is None:
        return False
    if pd.isna(valor):
        return False
    return bool(str(valor).strip())


def formatar_data_glpi(valor):
    if valor in (None, ""):
        return ""
    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y %H:%M:%S")
    texto = str(valor).strip()
    if not texto:
        return ""
    data = pd.to_datetime(texto, errors="coerce")
    if pd.isna(data):
        return texto
    return data.to_pydatetime().strftime("%d/%m/%Y %H:%M:%S")
