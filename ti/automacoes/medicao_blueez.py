import json
import os
import re
import shutil
import time
import traceback
import unicodedata
from decimal import Decimal, InvalidOperation
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from selenium import webdriver
from selenium.common.exceptions import ElementClickInterceptedException, TimeoutException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


BASE_URL = "https://app.blueez.com.br"
VERSAO_AUTOMACAO = "1.0.0"
COLUNAS_OBRIGATORIAS = [
    "empresa",
    "estabelecimento",
    "fornecedor",
    "cnpj_fornecedor",
    "contrato",
    "data_inicio",
    "data_fim",
    "competencia",
    "emissao_nf",
    "vencimento_nf",
    "id_nf",
    "item_medido",
    "valor_item",
    "observacao",
    "arquivo_1",
    "arquivo_2",
]


class AbortarAutomacao(Exception):
    pass


class ErroRegistroEsperado(Exception):
    pass


class ErroRegistroComSolicitacao(ErroRegistroEsperado):
    def __init__(self, mensagem, numero_solicitacao=""):
        super().__init__(mensagem)
        self.numero_solicitacao = numero_solicitacao


def executar(
    input_path=None,
    input_paths=None,
    attachments=None,
    output_dir=None,
    should_stop=None,
    log=None,
    parametros="",
    parametros_json=None,
):
    logger = log or (lambda message: None)
    logger(f"Medicao BlueEZ - Versao {VERSAO_AUTOMACAO}")
    inicio_rotina = datetime.now()
    inicio_monotonic = time.perf_counter()
    arquivos_principais = [Path(path) for path in (input_paths or ([] if input_path is None else [input_path]))]
    anexos = [Path(path) for path in (attachments or [])]
    pasta_saida = Path(output_dir or Path.cwd())
    pasta_saida.mkdir(parents=True, exist_ok=True)

    parametros_execucao = carregar_parametros(parametros, parametros_json)
    planilha_entrada = selecionar_planilha(arquivos_principais, input_path)
    planilha_trabalho = copiar_planilha(planilha_entrada, pasta_saida)
    logger(f"Planilha de trabalho criada: {planilha_trabalho.name}")
    logger(
        "Arquivos principais recebidos na execucao: "
        + (", ".join(path.name for path in arquivos_principais) if arquivos_principais else "nenhum")
    )
    logger(
        "Arquivos auxiliares recebidos na execucao: "
        + (", ".join(path.name for path in anexos) if anexos else "nenhum")
    )

    credenciais = resolver_credenciais(planilha_trabalho, parametros_execucao)
    df = pd.read_excel(planilha_trabalho)
    validar_planilha(df)
    numeros_solicitacao_conhecidos = carregar_numeros_solicitacao_existentes(df)
    attachment_index = indexar_anexos(anexos, logger)
    browser_timeout = int(parametros_execucao.get("browser_timeout", 40))
    pause_seconds = float(parametros_execucao.get("pause_seconds", 1))
    updates_excel = []
    updates_solicitacao = []
    resumo_execucao = {
        "ok": 0,
        "erro": 0,
        "teste": 0,
        "pulado_ok": 0,
        "tempo_total_registros_segundos": 0.0,
    }
    driver = None
    wait = None

    def checkpoint():
        if should_stop:
            should_stop()

    def pause(seconds=None):
        checkpoint()
        time.sleep(pause_seconds if seconds is None else seconds)

    try:
        logger(
            "Inicializando navegador Chrome para o BlueEZ "
            f"(headless={bool(parametros_execucao.get('headless', True))}, timeout={browser_timeout}s)."
        )
        driver = iniciar_driver(parametros_execucao)
        wait = WebDriverWait(driver, browser_timeout)
        logger(
            "Sessao Selenium criada com sucesso "
            f"(browser={safe_capability(driver, 'browserName')}, version={safe_capability(driver, 'browserVersion')})."
        )
        login_blueez(driver, wait, pause, credenciais, logger, parametros_execucao)
        for idx, registro in enumerate(df.to_dict(orient="records")):
            checkpoint()
            inicio_registro = time.perf_counter()
            linha_excel = idx + 2
            status_atual = ler_status(df, idx).strip().upper()
            if deve_pular_registro(status_atual):
                mensagem = f"Registro nao processado (pulado por status final: {status_atual})"
                resumo_execucao["pulado_ok"] += 1
                logger(f"Linha {linha_excel}: {mensagem}")
                continue

            try:
                logger(f"Iniciando registro {idx + 1}/{len(df)}.")
                resultado_registro = processar_registro(
                    driver,
                    wait,
                    pause,
                    registro,
                    attachment_index,
                    parametros_execucao,
                    logger,
                    numeros_solicitacao_conhecidos,
                )
                if resultado_registro["enviado"]:
                    numero_solicitacao = resultado_registro["numero_solicitacao"]
                    updates_excel.append((linha_excel, "OK", "Processado com sucesso"))
                    updates_solicitacao.append((linha_excel, numero_solicitacao))
                    numeros_solicitacao_conhecidos.add(numero_solicitacao)
                    resumo_execucao["ok"] += 1
                    logger(f"Linha {linha_excel}: OK | solicitacao={numero_solicitacao or 'nao capturada'}")
                else:
                    updates_excel.append((linha_excel, "TESTE", "Fluxo executado ate antes do envio final."))
                    updates_solicitacao.append((linha_excel, ""))
                    resumo_execucao["teste"] += 1
                    logger(f"Linha {linha_excel}: TESTE | envio final bloqueado propositalmente.")
            except ErroRegistroComSolicitacao as exc:
                mensagem = str(exc) or "Falha sem mensagem"
                updates_excel.append((linha_excel, "ERRO", mensagem))
                if exc.numero_solicitacao:
                    updates_solicitacao.append((linha_excel, exc.numero_solicitacao))
                    numeros_solicitacao_conhecidos.add(exc.numero_solicitacao)
                resumo_execucao["erro"] += 1
                logger(
                    f"Linha {linha_excel}: ERRO | {mensagem}"
                    + (
                        f" | solicitacao={exc.numero_solicitacao}"
                        if exc.numero_solicitacao
                        else ""
                    )
                )
            except ErroRegistroEsperado as exc:
                mensagem = str(exc) or "Falha sem mensagem"
                updates_excel.append((linha_excel, "ERRO", mensagem))
                resumo_execucao["erro"] += 1
                logger(f"Linha {linha_excel}: ERRO | {mensagem}")
            except Exception as exc:
                mensagem = str(exc) or "Falha sem mensagem"
                updates_excel.append((linha_excel, "ERRO", mensagem))
                resumo_execucao["erro"] += 1
                logger(f"Linha {linha_excel}: ERRO | {mensagem}")
                logger(traceback.format_exc().strip())
            finally:
                resumo_execucao["tempo_total_registros_segundos"] += max(
                    0.0,
                    time.perf_counter() - inicio_registro,
                )
    except Exception as exc:
        logger(f"Falha antes ou durante a fase inicial da automacao: {exc}")
        logger(traceback.format_exc().strip())
        if driver is not None:
            logger(
                "Estado do navegador no momento da falha inicial "
                f"(url_atual={safe_current_url(driver)}, titulo={safe_title(driver)!r})."
            )
        raise
    finally:
        try:
            if driver is not None:
                driver.quit()
        except Exception:
            pass

    fim_rotina = datetime.now()
    resumo_execucao["inicio_rotina"] = inicio_rotina.strftime("%d/%m/%Y %H:%M:%S")
    resumo_execucao["fim_rotina"] = fim_rotina.strftime("%d/%m/%Y %H:%M:%S")
    resumo_execucao["tempo_total_rotina_segundos"] = max(0.0, time.perf_counter() - inicio_monotonic)
    resumo_execucao["total_processado"] = (
        resumo_execucao.get("ok", 0)
        + resumo_execucao.get("erro", 0)
        + resumo_execucao.get("teste", 0)
    )
    resumo_execucao["tempo_medio_registro_segundos"] = (
        resumo_execucao["tempo_total_registros_segundos"] / resumo_execucao["total_processado"]
        if resumo_execucao["total_processado"]
        else 0.0
    )

    atualizar_status_mensagem_no_excel(planilha_trabalho, updates_excel)
    atualizar_numero_solicitacao_no_excel(planilha_trabalho, updates_solicitacao)
    salvar_resumo(
        pasta_saida,
        planilha_trabalho,
        updates_excel,
        updates_solicitacao,
        resumo_execucao,
    )
    logger(
        "Resumo final da execucao: "
        f"OK={resumo_execucao['ok']}, "
        f"ERRO={resumo_execucao['erro']}, "
        f"TESTE={resumo_execucao['teste']}, "
        f"PULADO_OK={resumo_execucao['pulado_ok']}, "
        f"TEMPO_TOTAL={formatar_duracao(resumo_execucao['tempo_total_rotina_segundos'])}, "
        f"TEMPO_MEDIO={formatar_duracao(resumo_execucao['tempo_medio_registro_segundos'])}."
    )
    modo_envio = "habilitado" if envio_final_habilitado(parametros_execucao) else "bloqueado"
    return {
        "message": (
            f"Medicao BlueEZ concluida. Arquivo gerado: {planilha_trabalho.name}. "
            f"Envio final estava {modo_envio}. "
            f"Resumo: OK={resumo_execucao['ok']}, ERRO={resumo_execucao['erro']}, "
            f"TESTE={resumo_execucao['teste']}, PULADO_OK={resumo_execucao['pulado_ok']}, "
            f"TEMPO_TOTAL={formatar_duracao(resumo_execucao['tempo_total_rotina_segundos'])}, "
            f"TEMPO_MEDIO={formatar_duracao(resumo_execucao['tempo_medio_registro_segundos'])}."
        )
    }


def carregar_parametros(parametros, parametros_json):
    if isinstance(parametros_json, dict):
        return dict(parametros_json)
    texto = (parametros or "").strip()
    if not texto:
        return {}
    try:
        valor = json.loads(texto)
        return valor if isinstance(valor, dict) else {"raw": valor}
    except json.JSONDecodeError:
        return {"raw": texto}


def resolver_credenciais(caminho_planilha, parametros_execucao):
    credenciais_planilha = ler_credenciais_da_planilha(caminho_planilha)
    if credenciais_planilha.get("username") and credenciais_planilha.get("password"):
        return credenciais_planilha

    username = parametros_execucao.get("username") or parametros_execucao.get("usuario") or os.getenv("BLUEEZ_USERNAME")
    password = parametros_execucao.get("password") or parametros_execucao.get("senha") or os.getenv("BLUEEZ_PASSWORD")
    if username and password:
        return {"username": username, "password": password}

    raise AbortarAutomacao(
        "Credenciais do BlueEZ ausentes. Coloque usuario e senha em uma aba separada da planilha principal "
        "ou informe username/password em parametros_texto (JSON) / variaveis de ambiente."
    )


def ler_credenciais_da_planilha(caminho_planilha):
    try:
        workbook = load_workbook(caminho_planilha, read_only=True, data_only=True)
    except Exception as exc:
        raise AbortarAutomacao(f"Nao foi possivel ler a planilha principal: {exc}") from exc

    sheet = None
    for nome in workbook.sheetnames:
        if nome.strip().lower() in {"credenciais", "login", "acesso", "usuario"}:
            sheet = workbook[nome]
            break

    if sheet is None:
        if len(workbook.sheetnames) >= 2:
            sheet = workbook[workbook.sheetnames[1]]
        else:
            return {}

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return {}

    headers = [normalizar_chave_excel(valor) for valor in header_row]
    values_row = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True), None)
    if not values_row:
        return {}

    data = {headers[idx]: values_row[idx] for idx in range(min(len(headers), len(values_row))) if headers[idx]}
    username = (
        data.get("usuario")
        or data.get("username")
        or data.get("login")
        or data.get("user")
    )
    password = (
        data.get("senha")
        or data.get("password")
        or data.get("pass")
    )
    return {
        "username": "" if username is None else str(username).strip(),
        "password": "" if password is None else str(password).strip(),
    }


def selecionar_planilha(arquivos_principais, input_path):
    candidatos = list(arquivos_principais)
    if input_path:
        candidatos.insert(0, Path(input_path))
    for arquivo in candidatos:
        if arquivo.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
            return arquivo
    raise AbortarAutomacao("Nenhuma planilha Excel foi enviada como arquivo principal.")


def copiar_planilha(planilha_entrada, pasta_saida):
    destino = pasta_saida / f"entrada_processada_{datetime.now().strftime('%Y%m%d_%H%M%S')}{planilha_entrada.suffix}"
    shutil.copy2(planilha_entrada, destino)
    return destino


def validar_planilha(df):
    faltando = [coluna for coluna in COLUNAS_OBRIGATORIAS if coluna not in df.columns]
    if faltando:
        raise AbortarAutomacao("Planilha invalida. Colunas ausentes: " + ", ".join(sorted(faltando)))


def carregar_numeros_solicitacao_existentes(df):
    numeros = set()
    for coluna in df.columns:
        if str(coluna).strip().lower() != "numero_solicitacao":
            continue
        for valor in df[coluna].tolist():
            if pd.isna(valor):
                continue
            numero = str(valor).strip()
            if numero:
                numeros.add(numero)
    return numeros


def deve_pular_registro(status_atual):
    return status_atual == "OK"




def normalizar_chave_excel(valor):
    texto = "" if valor is None else str(valor).strip().lower()
    texto = texto.replace(" ", "_")
    texto = re.sub(r"[^a-z0-9_]", "", texto)
    return texto


def iniciar_driver(parametros_execucao):
    options = Options()
    if bool(parametros_execucao.get("headless", True)):
        options.add_argument("--headless")
        options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--lang=pt-BR")
    options.add_argument("--accept-lang=pt-BR,pt")
    options.add_argument(f'--window-size={parametros_execucao.get("window_size", "1280,900")}')
    options.add_argument("--disable-extensions")
    options.add_argument("--no-first-run")
    options.add_argument("--no-default-browser-check")
    options.add_argument("--user-agent=Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")
    chrome_binary = parametros_execucao.get("chrome_binary")
    if chrome_binary:
        options.binary_location = chrome_binary
    elif os.name != "nt":
        options.binary_location = "/usr/bin/chromium"

    driver_path = parametros_execucao.get("driver_path")
    if not driver_path and os.name != "nt" and Path("/usr/bin/chromedriver").exists():
        driver_path = "/usr/bin/chromedriver"

    service = Service(executable_path=driver_path) if driver_path else Service()
    return webdriver.Chrome(service=service, options=options)


def login_blueez(driver, wait, pause, credenciais, logger, parametros_execucao):
    target_url = parametros_execucao.get("base_url", BASE_URL) + "/"
    logger(f"Abrindo portal BlueEZ em {target_url}")
    driver.get(target_url)
    pause()
    logger(
        "Pagina inicial carregada "
        f"(url_atual={safe_current_url(driver)}, titulo={safe_title(driver)!r})."
    )
    try:
        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'input[name="j_username"]')))
        logger("Formulario de login detectado. Preenchendo credenciais da planilha.")
    except TimeoutException:
        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#sidebarMenu, nav#sidebarMenu, .wrapperSidebarMenu")))
        logger(
            "Sessao reaproveitada detectada no BlueEZ "
            f"(url_atual={safe_current_url(driver)}, titulo={safe_title(driver)!r})."
        )
        return

    preencher_input(wait, 'input[name="j_username"]', credenciais["username"], pause)
    preencher_input(wait, 'input[name="j_password"]', credenciais["password"], pause)
    logger(f"Credenciais preenchidas para o usuario {credenciais['username']}. Enviando formulario.")
    driver.execute_script("const form = document.querySelector('form'); if (form) form.submit();")
    wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#sidebarMenu, nav#sidebarMenu, .wrapperSidebarMenu")))
    logger(
        f"Login realizado com sucesso para o usuario {credenciais['username']} "
        f"(url_atual={safe_current_url(driver)}, titulo={safe_title(driver)!r})."
    )


def processar_registro(
    driver,
    wait,
    pause,
    registro,
    attachment_index,
    parametros_execucao,
    logger,
    numeros_solicitacao_conhecidos,
):
    validar_campos_essenciais_registro(registro)
    logger("Abrindo tela inicial do BlueEZ para o registro atual.")
    driver.get(BASE_URL)
    pause()
    logger("Entrando no menu Medicao.")
    entrar_medicao(driver, wait, pause)
    numero_topo_anterior = capturar_numero_topo_grid_atual(driver)
    logger(
        "Ultimo numero visivel no grid antes do envio: "
        f"{numero_topo_anterior or 'nenhum registro visivel'}"
    )
    logger("Tela de Medicao aberta. Iniciando novo fluxo.")
    abrir_formulario_medicao(driver, wait, pause)
    logger("Formulario de medicao aberto com sucesso.")

    logger(f"Selecionando empresa: {str(registro['empresa']).strip()}")
    selecionar_modal(
        driver,
        wait,
        pause,
        'button[data-target="#modal_zoom_empresa_medicao_1"]',
        str(registro["empresa"]).strip(),
        "Empresa",
        "#id_descricao_empresa",
    )
    logger("Empresa selecionada com sucesso.")
    logger(f"Selecionando estabelecimento: {str(registro['estabelecimento']).strip()}")
    selecionar_modal(
        driver,
        wait,
        pause,
        'button[data-target="#modal_zoom_estabelecimento_medicao_1"]',
        str(registro["estabelecimento"]).strip(),
        "Estabelecimento",
        "#id_descricao_estabelecimento",
    )
    logger("Estabelecimento selecionado com sucesso.")
    logger(f"Selecionando fornecedor pelo CNPJ: {registro['cnpj_fornecedor']}")
    selecionar_fornecedor(driver, wait, pause, registro["cnpj_fornecedor"])
    logger("Fornecedor selecionado com sucesso.")
    logger(f"Selecionando contrato: {registro['contrato']}")
    selecionar_contrato(driver, wait, pause, registro["contrato"])
    logger("Contrato selecionado com sucesso.")
    logger("Preenchendo datas da medicao.")
    preencher_datas(driver, wait, pause, registro)
    logger(
        "Datas preenchidas com sucesso "
        f"(inicio={formatar_data(registro['data_inicio'])}, "
        f"fim={formatar_data(registro['data_fim'])}, "
        f"competencia={formatar_competencia(registro['competencia'])})."
    )
    logger(f"Preenchendo numero da NF: {registro['id_nf']}")
    preencher_input_tab(driver, wait, "#id_nf_medicao", str(registro["id_nf"]), pause)
    logger("Numero da NF preenchido com sucesso.")
    logger(
        "Preenchendo item medido "
        f"(item={registro['item_medido']}, valor={registro['valor_item']})."
    )
    preencher_item_medido(
        driver,
        wait,
        pause,
        registro["item_medido"],
        registro["valor_item"],
        logger,
    )
    logger("Item medido preenchido com sucesso.")
    logger("Preenchendo observacao.")
    preencher_observacao(wait, pause, registro["observacao"])
    logger("Observacao preenchida com sucesso.")
    observacao_aplicada = capturar_valor_observacao(wait)
    logger(f"Observacao registrada no BlueEZ: {observacao_aplicada or 'vazio'}")
    nomes_esperados = listar_nomes_anexos_registro(registro)
    if nomes_esperados:
        logger("Resolvendo anexos do registro: " + ", ".join(nomes_esperados))
    else:
        logger("Registro sem anexos informados na planilha.")
    arquivos_upload = resolver_arquivos_registro(registro, attachment_index, logger)
    if arquivos_upload:
        logger(
            "Anexos localizados para upload: "
            + ", ".join(arquivo.name for arquivo in arquivos_upload)
        )
    fazer_upload(driver, wait, pause, arquivos_upload, logger)
    if not envio_final_habilitado(parametros_execucao):
        logger("Teste seguro: clique final de envio bloqueado por configuracao.")
        return {"enviado": False, "numero_solicitacao": ""}
    logger("Clique final habilitado. Enviando solicitacao.")
    clicar_enviar(driver, wait, pause)
    logger("Solicitacao enviada. Validando criacao do novo registro no grid.")
    numero_solicitacao = validar_e_capturar_nova_solicitacao(
        driver,
        wait,
        pause,
        numero_topo_anterior,
        numeros_solicitacao_conhecidos,
        logger,
    )
    return {"enviado": True, "numero_solicitacao": numero_solicitacao}


def entrar_medicao(driver, wait, pause):
    driver.switch_to.default_content()
    aguardar_shell_blueez(driver, wait)
    remover_overlays_interativos(driver)
    dropdown = localizar_dropdown_medicao(driver)
    submenu_visivel = driver.execute_script(
        "const el = document.querySelector('#pageSubmenu30'); return !!(el && el.offsetParent !== null);"
    )
    if not submenu_visivel:
        driver.execute_script("arguments[0].click();", dropdown)
        pause()
    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#pageSubmenu30")))
    wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="pageSubmenu30"]//span[normalize-space()="Medição"]'))).click()
    reentrar_iframe(driver, wait, pause)
    pause()


def reentrar_iframe(driver, wait, pause):
    driver.switch_to.default_content()
    aguardar_shell_blueez(driver, wait)
    wait.until(EC.frame_to_be_available_and_switch_to_it((By.CSS_SELECTOR, "iframe#iframeContent")))
    pause()


def aguardar_shell_blueez(driver, wait):
    wait.until(lambda d: d.execute_script("return document.readyState") == "complete")
    wait.until(
        EC.presence_of_element_located(
            (By.CSS_SELECTOR, "#sidebarMenu, nav#sidebarMenu, .wrapperSidebarMenu")
        )
    )


def localizar_dropdown_medicao(driver):
    for seletor in [
        'a.dropdownSidebar-toggle[aria-controls="pageSubmenu30"]',
        '[aria-controls="pageSubmenu30"]',
        'a[href="#pageSubmenu30"]',
    ]:
        elementos = driver.find_elements(By.CSS_SELECTOR, seletor)
        for elemento in elementos:
            try:
                if elemento.is_displayed():
                    return elemento
            except Exception:
                continue
    raise TimeoutException(
        "Menu lateral do BlueEZ nao ficou disponivel para abrir Medicao. "
        f"url_atual={safe_current_url(driver)} | titulo={safe_title(driver)!r}"
    )


def abrir_formulario_medicao(driver, wait, pause):
    reentrar_iframe(driver, wait, pause)
    botao_novo = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#new_flow")))
    driver.execute_script("arguments[0].click();", botao_novo)
    pause()
    reentrar_iframe(driver, wait, pause)
    wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, ".modal.fade.show, .modal.show")))
    wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#id_data_inicio")))


def selecionar_modal(driver, wait, pause, seletor_botao, termo_busca, label, seletor_resultado=None):
    botao = obter_elemento_clicavel(
        driver,
        wait,
        seletor_botao,
        f"botao de selecao de {label}",
    )
    driver.execute_script("arguments[0].click();", botao)
    pause()
    modal_id = botao.get_attribute("data-target") or ""
    modal_selector = f"{modal_id} input.form-control[placeholder=\"Procurar\"]" if modal_id else 'input.form-control[placeholder="Procurar"]'
    campo = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, modal_selector)))
    campo.clear()
    campo.send_keys(termo_busca)
    campo.send_keys(Keys.ENTER)
    pause()

    linhas = []
    if modal_id:
        linhas = [
            linha
            for linha in driver.find_elements(By.CSS_SELECTOR, f"{modal_id} table tbody tr")
            if linha.is_displayed() and termo_busca.lower() in normalizar_texto(linha.text)
        ]
    if not linhas:
        sugestoes = coletar_amostra_modal(driver, modal_id)
        raise ValueError(montar_erro_busca(label, termo_busca, sugestoes))

    driver.execute_script("arguments[0].click();", linhas[0])
    pause()
    if seletor_resultado:
        wait.until(
            lambda d: termo_busca.lower() in normalizar_texto(
                d.find_element(By.CSS_SELECTOR, seletor_resultado).get_attribute("value") or ""
            )
        )


def selecionar_fornecedor(driver, wait, pause, cnpj):
    botao = obter_elemento_clicavel(
        driver,
        wait,
        'button[data-target="#modal_zoom_fornecedor_medicao_1"]',
        "botao de selecao de Fornecedor",
    )
    driver.execute_script("arguments[0].click();", botao)
    pause()
    cnpj_formatado = formatar_cnpj(cnpj)
    campo = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, '#tabela_zoom_fornecedor_medicao_1_wrapper input.form-control[placeholder="Procurar"]')))
    campo.clear()
    campo.send_keys(cnpj_formatado)
    campo.send_keys(Keys.ENTER)
    pause(2)
    linhas = driver.find_elements(
        By.XPATH,
        f'//*[@id="tabela_zoom_fornecedor_medicao_1"]//tbody/tr[td[1][normalize-space()="{cnpj_formatado}"]]'
    )
    if not linhas:
        sugestoes = coletar_amostra_tabela(driver, "#tabela_zoom_fornecedor_medicao_1 tbody tr")
        raise ValueError(
            montar_erro_busca(
                "Fornecedor",
                cnpj_formatado,
                sugestoes,
                detalhe_extra="Verifique se o CNPJ da planilha existe e esta vinculado no BlueEZ.",
            )
        )
    linha = wait.until(EC.visibility_of(linhas[0]))
    driver.execute_script("arguments[0].click();", linha)
    pause()


def selecionar_contrato(driver, wait, pause, contrato):
    botao = obter_elemento_clicavel(
        driver,
        wait,
        'button[data-target="#modal_zoom_contrato_medicao_1"]',
        "botao de selecao de Contrato",
    )
    driver.execute_script("arguments[0].click();", botao)
    pause()
    numero_contrato = limpar_codigo_excel(contrato)
    seletores_busca = [
        '#tabela_zoom_contrato_medicao_1_wrapper input.form-control[placeholder="Procurar"]',
        '#modal_zoom_contrato_medicao_1 input.form-control[placeholder="Procurar"]',
        '#modal_zoom_contrato_medicao_1 input[type="text"]',
    ]
    campo = None
    for seletor in seletores_busca:
        elementos = driver.find_elements(By.CSS_SELECTOR, seletor)
        elementos_visiveis = [elemento for elemento in elementos if elemento.is_displayed() and elemento.is_enabled()]
        if elementos_visiveis:
            campo = elementos_visiveis[0]
            break
    if campo is None:
        raise ValueError(
            "Campo de busca do Contrato nao apareceu no BlueEZ apos abrir o modal. "
            f"url_atual={safe_current_url(driver)} | titulo={safe_title(driver)!r}"
        )
    campo.clear()
    campo.send_keys(numero_contrato)
    campo.send_keys(Keys.ENTER)
    pause()
    linhas = [
        linha
        for linha in driver.find_elements(By.CSS_SELECTOR, "#tabela_zoom_contrato_medicao_1 tbody tr")
        if linha.is_displayed()
    ]
    if not linhas:
        sugestoes = coletar_amostra_tabela(driver, "#tabela_zoom_contrato_medicao_1 tbody tr")
        raise ValueError(
            montar_erro_busca(
                "Contrato",
                numero_contrato,
                sugestoes,
                detalhe_extra="Confirme se o numero do contrato esta correto e disponivel para o usuario logado.",
            )
        )
    driver.execute_script("arguments[0].click();", linhas[0])
    pause()


def preencher_datas(driver, wait, pause, registro):
    for seletor, valor in {
        "#id_data_inicio": formatar_data(registro["data_inicio"]),
        "#id_data_fim": formatar_data(registro["data_fim"]),
        "#id_competencia": formatar_competencia(registro["competencia"]),
        "#id_data_emissao": formatar_data(registro["emissao_nf"]),
        "#id_data_vencimento": formatar_data(registro["vencimento_nf"]),
    }.items():
        preencher_input_tab(driver, wait, seletor, valor, pause)


def preencher_item_medido(driver, wait, pause, item_medido, valor_item, logger=None):
    campos = driver.find_elements(By.CSS_SELECTOR, 'div[class^="linha_valor_"] input.form-control.input-blueez')
    indice = int(float(item_medido)) - 1
    if indice < 0 or indice >= len(campos):
        raise ValueError(f"item_medido invalido: {item_medido}")
    valor = f"{float(valor_item):.2f}".replace(".", ",")
    for posicao, campo in enumerate(campos):
        valor_esperado = valor if posicao == indice else "0,00"
        valor_lido = preencher_campo_monetario(
            driver,
            campo,
            valor_esperado,
            pause,
            logger,
        )
        if logger:
            logger(
                f"Auditoria item {posicao + 1}: esperado={valor_esperado} | exibido={valor_lido or 'vazio'}"
            )


def preencher_observacao(wait, pause, observacao):
    campo = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#id_observacao_medicao")))
    campo.clear()
    pause()
    campo.send_keys("" if pd.isna(observacao) else str(observacao))
    pause()


def capturar_valor_observacao(wait):
    campo = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#id_observacao_medicao")))
    return (campo.get_attribute("value") or "").strip()


def capturar_valores_itens(driver):
    campos = driver.find_elements(By.CSS_SELECTOR, 'div[class^="linha_valor_"] input.form-control.input-blueez')
    return [(campo.get_attribute("value") or "").strip() for campo in campos]


def fazer_upload(driver, wait, pause, arquivos, logger=None):
    if not arquivos:
        if logger:
            logger("Nenhum anexo para enviar neste registro.")
        return
    if logger:
        logger("Abrindo aba de anexos.")
    clicar_elemento_seguro(
        driver,
        wait,
        (By.CSS_SELECTOR, 'a[href="#content_tab_1_50"]'),
        pause,
        logger,
        "aba de anexos",
    )
    pause()
    if logger:
        logger("Aba de anexos aberta. Clicando em novo anexo.")
    clicar_elemento_seguro(
        driver,
        wait,
        (By.CSS_SELECTOR, "button.btn-orange-componente"),
        pause,
        logger,
        "botao novo anexo",
    )
    pause()
    if logger:
        logger("Campo de upload localizado. Enviando arquivos para o navegador.")
    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#upload-file"))).send_keys("\n".join(str(arquivo) for arquivo in arquivos))
    pause()
    if logger:
        logger("Arquivos enviados ao input. Acionando upload no BlueEZ.")
    clicar_elemento_seguro(
        driver,
        wait,
        (By.CSS_SELECTOR, "button.btn-upload-files"),
        pause,
        logger,
        "botao upload",
    )
    pause(2)
    if logger:
        logger("Comando de upload executado. Prosseguindo para a proxima etapa.")


def clicar_enviar(driver, wait, pause):
    driver.switch_to.default_content()
    pause()
    try:
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "iframe#iframeContent")))
        driver.switch_to.frame(driver.find_element(By.CSS_SELECTOR, "iframe#iframeContent"))
    except Exception:
        pass
    for seletor in ["#botao_enviar_solicitacao_0", 'button[id^="botao_enviar_solicitacao_"]', "button.btn-success", 'button[type="submit"]']:
        botoes = driver.find_elements(By.CSS_SELECTOR, seletor)
        for botao in botoes:
            if botao.is_displayed() and botao.is_enabled():
                driver.execute_script("arguments[0].click();", botao)
                pause(2)
                return
    raise TimeoutException("Botao de enviar nao encontrado no contexto atual.")


def capturar_numero_solicitacao(driver, wait, pause):
    try:
        driver.switch_to.default_content()
    except Exception:
        return ""

    pause(0.2)
    iframes = driver.find_elements(By.CSS_SELECTOR, "iframe#iframeContent")
    if not iframes:
        return ""

    try:
        driver.switch_to.frame(iframes[0])
    except Exception:
        return ""

    remover_overlays_interativos(driver)
    linhas = driver.find_elements(By.CSS_SELECTOR, "#DataTables_Table_0 tbody tr")
    for linha in linhas:
        try:
            if not linha.is_displayed():
                continue
            colunas = linha.find_elements(By.TAG_NAME, "td")
            if not colunas:
                continue
            numero = (colunas[0].text or "").strip()
            if numero and numero.isdigit():
                return numero
        except Exception:
            continue
    return ""


def capturar_primeira_linha_grid(driver):
    try:
        linhas = driver.find_elements(By.CSS_SELECTOR, "#DataTables_Table_0 tbody tr")
        for linha in linhas:
            if not linha.is_displayed():
                continue
            colunas = linha.find_elements(By.TAG_NAME, "td")
            if not colunas:
                continue
            return [(coluna.text or "").strip() for coluna in colunas]
    except Exception:
        return []
    return []


def capturar_detalhes_primeiro_registro_grid(driver, wait, pause):
    try:
        driver.switch_to.default_content()
    except Exception:
        return {}

    pause(0.2)
    iframes = driver.find_elements(By.CSS_SELECTOR, "iframe#iframeContent")
    if not iframes:
        return {}

    try:
        driver.switch_to.frame(iframes[0])
    except Exception:
        return {}

    remover_overlays_interativos(driver)
    colunas = capturar_primeira_linha_grid(driver)
    if not colunas:
        return {}

    return {
        "numero": colunas[0] if len(colunas) > 0 else "",
        "solicitante": colunas[1] if len(colunas) > 1 else "",
        "etapa": colunas[2] if len(colunas) > 2 else "",
        "responsavel": colunas[3] if len(colunas) > 3 else "",
        "data_criacao": colunas[4] if len(colunas) > 4 else "",
    }


def capturar_numero_topo_grid_atual(driver):
    try:
        linhas = driver.find_elements(By.CSS_SELECTOR, "#DataTables_Table_0 tbody tr")
        for linha in linhas:
            if not linha.is_displayed():
                continue
            colunas = linha.find_elements(By.TAG_NAME, "td")
            if not colunas:
                continue
            numero = (colunas[0].text or "").strip()
            if numero:
                return numero
    except Exception:
        return ""
    return ""


def validar_e_capturar_nova_solicitacao(
    driver,
    wait,
    pause,
    numero_topo_anterior,
    numeros_solicitacao_conhecidos,
    logger,
    timeout=25,
):
    fim = time.time() + timeout
    ultimo_numero_lido = ""

    while time.time() < fim:
        numero_atual = capturar_numero_solicitacao(driver, wait, pause)
        ultimo_numero_lido = numero_atual

        if not numero_atual:
            logger(
                "Aguardando atualizacao do grid apos o envio. "
                + diagnostico_pos_envio(driver)
            )
            time.sleep(1)
            continue

        if numero_topo_anterior and numero_atual == numero_topo_anterior:
            logger(
                "Grid ainda exibe o numero anterior apos o envio. "
                f"numero_atual={numero_atual}. "
                + diagnostico_pos_envio(driver)
            )
            time.sleep(1)
            continue

        if numero_atual in numeros_solicitacao_conhecidos:
            raise ValueError(
                "O BlueEZ retornou um numero de solicitacao que ja existe na planilha/execucao: "
                f"{numero_atual}. Isso indica que o novo registro provavelmente nao foi criado."
            )

        detalhes_registro = capturar_detalhes_primeiro_registro_grid(driver, wait, pause)
        etapa = normalizar_texto(detalhes_registro.get("etapa"))
        if detalhes_registro.get("numero") == numero_atual and "inconsist" in etapa:
            responsavel = detalhes_registro.get("responsavel") or "Responsavel nao identificado no BlueEZ"
            raise ErroRegistroComSolicitacao(
                "A medicao foi criada no BlueEZ, mas ficou com etapa 'Inconsistencia'. "
                "Verifique o contrato informado e entre em contato com o responsavel pelos contratos e medicoes no BlueEZ. "
                f"Solicitacao gerada: {numero_atual}. Responsavel atual no BlueEZ: {responsavel}.",
                numero_solicitacao=numero_atual,
            )

        logger(f"Novo numero de solicitacao capturado com sucesso: {numero_atual}")
        return numero_atual

    if numero_topo_anterior and ultimo_numero_lido == numero_topo_anterior:
        raise ValueError(
            "O BlueEZ voltou para o grid, mas o ultimo numero visivel permaneceu igual ao registro anterior "
            f"({numero_topo_anterior}). Isso indica que a medicao atual provavelmente nao foi salva."
        )

    raise ValueError(
        "Nao foi possivel confirmar a criacao de um novo registro no grid do BlueEZ apos o envio. "
        + diagnostico_pos_envio(driver)
    )


def diagnostico_pos_envio(driver):
    try:
        driver.switch_to.default_content()
    except Exception:
        return "Diagnostico indisponivel."

    toasts = []
    for seletor in [
        ".toast-error",
        ".toast-success",
        ".alert-danger",
        ".alert-success",
        "div.toast",
        'div[role=\"alert\"]',
        ".swal2-popup",
    ]:
        for el in driver.find_elements(By.CSS_SELECTOR, seletor):
            try:
                if el.is_displayed():
                    texto = (el.text or "").strip()
                    if texto and texto not in toasts:
                        toasts.append(texto)
            except Exception:
                continue

    iframes = driver.find_elements(By.CSS_SELECTOR, "iframe#iframeContent")
    if not iframes:
        return "Nenhum iframe de conteudo foi encontrado apos o envio."

    try:
        driver.switch_to.frame(iframes[0])
    except Exception:
        return "O iframe de conteudo existe, mas nao foi possivel acessa-lo apos o envio."

    remover_overlays_interativos(driver)
    linhas = driver.find_elements(By.CSS_SELECTOR, "#DataTables_Table_0 tbody tr")
    grid_count = sum(1 for linha in linhas if linha.is_displayed())
    primeiro_numero = ""
    for linha in linhas:
        try:
            if not linha.is_displayed():
                continue
            colunas = linha.find_elements(By.TAG_NAME, "td")
            if colunas:
                primeiro_numero = (colunas[0].text or "").strip()
                break
        except Exception:
            continue

    modais_visiveis = 0
    for modal in driver.find_elements(By.CSS_SELECTOR, ".modal.fade.show, .modal.show"):
        try:
            if modal.is_displayed():
                modais_visiveis += 1
        except Exception:
            continue

    partes = [
        f"toast={' | '.join(toasts) if toasts else 'nenhum'}",
        f"grid_linhas={grid_count}",
        f"primeiro_numero={primeiro_numero or 'nenhum'}",
        f"modais_visiveis={modais_visiveis}",
    ]
    return "Diagnostico pos-envio: " + ", ".join(partes) + "."


def preencher_input(wait, seletor, valor, pause):
    campo = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, seletor)))
    campo.clear()
    pause()
    campo.send_keys(valor)
    pause()


def preencher_input_tab(driver, wait, seletor, valor, pause):
    campo = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, seletor)))
    preencher_elemento_tab(driver, campo, valor, pause)


def preencher_campo_monetario(driver, campo, valor, pause, logger=None):
    preparar_campo_para_digitacao(driver, campo)
    pause()
    input_type = (campo.get_attribute("type") or "").strip().lower()
    if input_type == "number":
        valor_number = valor_para_input_number(valor)
        tentativas = [
            ("number_com_ponto", valor_number, False),
            ("number_set_value", valor_number, True),
        ]
    else:
        tentativas = [
            ("literal_com_virgula", valor, False),
            ("mascara_por_digitos", digitos_monetarios(valor), False),
            ("set_value_nativo", valor, True),
        ]

    for nome_tentativa, valor_digitado, usar_set_value in tentativas:
        focar_campo(driver, campo)
        pause()
        limpar_campo_monetario(driver, campo, pause)
        pause()

        if usar_set_value:
            set_value_native(driver, campo, valor_digitado)
            pause()
        else:
            campo.send_keys(valor_digitado)
            pause()

        driver.execute_script(
            """
            arguments[0].dispatchEvent(new KeyboardEvent('keyup', { bubbles: true }));
            arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
            arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
            arguments[0].dispatchEvent(new Event('blur', { bubbles: true }));
            """,
            campo,
        )
        pause()
        try:
            campo.send_keys(Keys.TAB)
            pause()
        except Exception:
            pass

        valor_tela = (campo.get_attribute("value") or "").strip()
        if logger:
            logger(
                f"Auditoria campo monetario ({nome_tentativa}): enviado={valor_digitado} | exibido={valor_tela or 'vazio'}"
            )
        if valor_monetario_valido_na_tela(valor_tela, valor):
            return valor_tela

    raise ValueError(
        "O campo monetario do item nao refletiu o valor esperado no BlueEZ. "
        f"Esperado: {valor}. Exibido na tela: {valor_tela or 'vazio'}."
    )


def preencher_elemento_tab(driver, campo, valor, pause):
    preparar_campo_para_digitacao(driver, campo)
    pause()
    set_value_native(driver, campo, valor)
    pause()

    try:
        driver.execute_script("arguments[0].focus();", campo)
    except Exception:
        pass

    try:
        campo.send_keys(Keys.CONTROL, "a")
        pause()
        campo.send_keys(valor)
        pause()
        campo.send_keys(Keys.TAB)
        pause()
    except Exception:
        driver.execute_script("arguments[0].dispatchEvent(new Event('blur', { bubbles: true }));", campo)
        pause()


def preparar_campo_para_digitacao(driver, campo):
    remover_overlays_interativos(driver)
    driver.execute_script(
        """
        const el = arguments[0];
        const footer = document.querySelector('#icones-footer');
        if (footer) {
            footer.style.pointerEvents = 'none';
            footer.style.opacity = '0';
        }
        el.scrollIntoView({block: 'center', inline: 'nearest'});
        """,
        campo,
    )


def focar_campo(driver, campo):
    driver.execute_script(
        """
        arguments[0].focus();
        arguments[0].click();
        """,
        campo,
    )


def remover_overlays_interativos(driver):
    driver.execute_script(
        """
        const selectors = [
            '#icones-footer',
            '#nps-toast',
            '.nps-toast',
            '.nps-container',
            '.nps-overlay',
            '[class*="nps"]'
        ];
        selectors.forEach((selector) => {
            document.querySelectorAll(selector).forEach((el) => {
                el.style.pointerEvents = 'none';
                el.style.opacity = '0';
                el.style.visibility = 'hidden';
            });
        });
        """
    )


def clicar_elemento_seguro(driver, wait, locator, pause, logger=None, descricao="elemento"):
    ultimo_erro = None
    for tentativa in range(1, 4):
        pause()
        remover_overlays_interativos(driver)
        elemento = wait.until(EC.presence_of_element_located(locator))
        driver.execute_script(
            "arguments[0].scrollIntoView({block:'center', inline:'nearest'});",
            elemento,
        )
        pause(0.3)
        try:
            wait.until(EC.element_to_be_clickable(locator)).click()
            return
        except ElementClickInterceptedException as exc:
            ultimo_erro = exc
            remover_overlays_interativos(driver)
            pause(0.2)
            try:
                driver.execute_script("arguments[0].click();", elemento)
                return
            except Exception as js_exc:
                ultimo_erro = js_exc
                if logger:
                    logger(
                        f"Clique interceptado em {descricao} na tentativa {tentativa}/3. "
                        "Aplicando limpeza de overlay e retry."
                    )
                pause(0.5)
        except Exception as exc:
            ultimo_erro = exc
            remover_overlays_interativos(driver)
            pause(0.3)
            try:
                driver.execute_script("arguments[0].click();", elemento)
                return
            except Exception as js_exc:
                ultimo_erro = js_exc
                pause(0.5)
    raise ultimo_erro


def set_value_native(driver, campo, valor):
    driver.execute_script(
        """
        const el = arguments[0];
        const value = arguments[1];
        el.removeAttribute('readonly');
        el.removeAttribute('disabled');
        const proto = el instanceof HTMLInputElement ? HTMLInputElement.prototype : el.__proto__;
        const descriptor = Object.getOwnPropertyDescriptor(proto, 'value');
        if (descriptor && descriptor.set) {
            descriptor.set.call(el, value);
        } else {
            el.value = value;
        }
        el.dispatchEvent(new Event('input', { bubbles: true }));
        el.dispatchEvent(new Event('change', { bubbles: true }));
        """,
        campo,
        valor,
    )


def limpar_campo(driver, campo):
    driver.execute_script(
        """
        arguments[0].focus();
        arguments[0].value = '';
        arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
        arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
        """,
        campo,
    )


def limpar_campo_monetario(driver, campo, pause):
    limpar_campo(driver, campo)
    pause()
    try:
        campo.send_keys(Keys.CONTROL, "a")
        pause()
        campo.send_keys(Keys.BACKSPACE)
        pause()
        campo.send_keys(Keys.DELETE)
        pause()
        valor_tela = (campo.get_attribute("value") or "").strip()
        if valor_tela in {"0", "0,00", "0.00"}:
            campo.send_keys(Keys.BACKSPACE)
            pause()
            campo.send_keys(Keys.DELETE)
            pause()
    except Exception:
        pass


def parse_decimal_br(valor):
    texto = str(valor or "").strip()
    if not texto:
        raise InvalidOperation("valor vazio")

    texto = texto.replace("R$", "").replace(" ", "")
    texto = re.sub(r"[^0-9,.\-]", "", texto)
    if not texto:
        raise InvalidOperation("valor invalido")

    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    elif texto.count(".") > 1:
        texto = texto.replace(".", "")

    return Decimal(texto).quantize(Decimal("0.01"))


def digitos_monetarios(valor):
    decimal = parse_decimal_br(valor)
    centavos = int((decimal * 100).quantize(Decimal("1")))
    return str(centavos)


def valor_para_input_number(valor):
    decimal = parse_decimal_br(valor)
    if decimal == Decimal("0.00"):
        return "0"
    return format(decimal, ".2f")


def valores_monetarios_equivalentes(valor_tela, valor_esperado):
    try:
        return parse_decimal_br(valor_tela) == parse_decimal_br(valor_esperado)
    except (InvalidOperation, ValueError):
        return False


def valor_monetario_valido_na_tela(valor_tela, valor_esperado):
    valor_tela = str(valor_tela or "").strip()

    if not valor_tela:
        return False

    return valores_monetarios_equivalentes(valor_tela, valor_esperado)


def indexar_anexos(anexos, logger=None):
    indexed = {}
    for anexo in anexos:
        anexo_normalizado = normalizar_extensao_minuscula(anexo, logger)
        indexed[anexo_normalizado.name.lower()] = anexo_normalizado
    return indexed


def resolver_arquivos_registro(registro, attachment_index, logger=None):
    arquivos = []
    anexos_disponiveis = listar_anexos_disponiveis(attachment_index)
    sem_anexos_enviados = not attachment_index
    for chave in ["arquivo_1", "arquivo_2"]:
        nome = registro.get(chave)
        if not isinstance(nome, str) or not nome.strip():
            continue
        nome_limpo = nome.strip()
        if logger:
            logger(f"Procurando anexo informado na planilha: {nome_limpo}")
        encontrado = localizar_anexo(nome_limpo, attachment_index, logger)
        if not encontrado:
            if sem_anexos_enviados:
                raise ErroRegistroEsperado(
                    "A linha exige arquivo auxiliar, mas nenhum anexo foi enviado nesta execucao. "
                    f"Arquivo esperado pela planilha: {nome_limpo}. "
                    "Envie os arquivos no campo 'Arquivos auxiliares' para processar esta linha."
                )
            raise ErroRegistroEsperado(
                "Arquivo auxiliar nao localizado para esta linha. "
                f"Arquivo esperado pela planilha: {nome_limpo}. "
                f"Anexos recebidos na execucao: {anexos_disponiveis}"
            )
        arquivos.append(encontrado)
    return arquivos


def localizar_anexo(nome, attachment_index, logger=None):
    chave = nome.lower()
    if chave in attachment_index:
        if logger:
            logger(f"Anexo localizado por nome exato: {attachment_index[chave].name}")
        return attachment_index[chave]

    base, extensao = os.path.splitext(nome)
    if not extensao:
        chave_pdf = f"{base}.pdf".lower()
        if chave_pdf in attachment_index:
            if logger:
                logger(
                    "Anexo localizado assumindo extensao .pdf ausente na planilha: "
                    f"{attachment_index[chave_pdf].name}"
                )
            return attachment_index[chave_pdf]

    nome_normalizado = normalizar_nome_anexo(nome)
    stem_normalizado = normalizar_nome_anexo(base or nome)

    candidatos_nome = []
    candidatos_stem = []
    candidatos_parciais = []
    for caminho in attachment_index.values():
        nome_arquivo_normalizado = normalizar_nome_anexo(caminho.name)
        stem_arquivo_normalizado = normalizar_nome_anexo(caminho.stem)

        if nome_normalizado and nome_normalizado == nome_arquivo_normalizado:
            candidatos_nome.append(caminho)
            continue

        if stem_normalizado and stem_normalizado == stem_arquivo_normalizado:
            candidatos_stem.append(caminho)
            continue

        pontuacao = pontuar_correspondencia_anexo(stem_normalizado, stem_arquivo_normalizado)
        if pontuacao > 0:
            candidatos_parciais.append((pontuacao, caminho))

    if len(candidatos_nome) == 1:
        if logger:
            logger(
                "Anexo localizado por normalizacao de nome "
                f"(acentos/espacos/caracteres): {candidatos_nome[0].name}"
            )
        return candidatos_nome[0]

    if len(candidatos_stem) == 1:
        if logger:
            logger(
                "Anexo localizado por nome base, ignorando extensao: "
                f"{candidatos_stem[0].name}"
            )
        return candidatos_stem[0]

    if candidatos_parciais:
        candidatos_parciais.sort(
            key=lambda item: (
                item[0],
                len(normalizar_nome_anexo(item[1].stem)),
            ),
            reverse=True,
        )
        melhor_pontuacao, melhor_caminho = candidatos_parciais[0]
        if len(candidatos_parciais) == 1 or (
            len(candidatos_parciais) > 1 and melhor_pontuacao > candidatos_parciais[1][0]
        ):
            if logger:
                logger(
                    "Anexo localizado por correspondencia parcial inteligente: "
                    f"{melhor_caminho.name} (score={melhor_pontuacao})"
                )
            return melhor_caminho
    return None


def listar_nomes_anexos_registro(registro):
    nomes = []
    for chave in ["arquivo_1", "arquivo_2"]:
        valor = registro.get(chave)
        if isinstance(valor, str) and valor.strip():
            nomes.append(valor.strip())
    return nomes


def listar_anexos_disponiveis(attachment_index):
    nomes = sorted(caminho.name for caminho in attachment_index.values())
    return ", ".join(nomes) if nomes else "nenhum arquivo auxiliar foi enviado"


def normalizar_extensao_minuscula(anexo, logger=None):
    sufixo = anexo.suffix
    if not sufixo or sufixo == sufixo.lower():
        return anexo

    destino = anexo.with_name(f"{anexo.stem}{sufixo.lower()}")
    if destino == anexo:
        return anexo

    if destino.exists():
        if logger:
            logger(
                "Extensao em maiuscula detectada, mas ja existe arquivo com extensao minuscula. "
                f"Usando {destino.name} no lugar de {anexo.name}."
            )
        return destino

    anexo.rename(destino)
    if logger:
        logger(
            "Extensao do anexo ajustada para minuscula: "
            f"{anexo.name} -> {destino.name}"
        )
    return destino


def normalizar_nome_anexo(valor):
    texto = normalizar_texto(valor)
    # Cola separadores monetarios/milhar entre digitos para compatibilizar
    # nomes vindos da planilha com anexos sanitizados pelo upload.
    texto = re.sub(r"(?<=\d)[\s.,_-]+(?=\d)", "", texto)
    texto = re.sub(r"\s+", " ", texto).strip()
    texto = re.sub(r"[^a-z0-9]+", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


def pontuar_correspondencia_anexo(nome_buscado, nome_candidato):
    if not nome_buscado or not nome_candidato:
        return 0
    if nome_buscado == nome_candidato:
        return 1000
    if nome_buscado in nome_candidato:
        return 900 - abs(len(nome_candidato) - len(nome_buscado))

    tokens_buscados = [token for token in nome_buscado.split() if token]
    tokens_candidato = [token for token in nome_candidato.split() if token]
    if not tokens_buscados or not tokens_candidato:
        return 0

    conjunto_candidato = set(tokens_candidato)
    intersecao = [token for token in tokens_buscados if token in conjunto_candidato]
    if not intersecao:
        return 0

    if all(token in conjunto_candidato for token in tokens_buscados):
        return 800 - abs(len(tokens_candidato) - len(tokens_buscados))

    score = len(intersecao) * 100
    if tokens_buscados[0] in conjunto_candidato:
        score += 25
    if tokens_buscados[-1] in conjunto_candidato:
        score += 25
    return score if len(intersecao) >= max(2, len(tokens_buscados) - 1) else 0


def valor_planilha_preenchido(valor):
    if valor is None or pd.isna(valor):
        return False
    texto = str(valor).strip()
    return bool(texto and texto.lower() != "nan")


def validar_campos_essenciais_registro(registro):
    campos_essenciais = {
        "empresa": "Empresa",
        "estabelecimento": "Estabelecimento",
        "cnpj_fornecedor": "CNPJ do fornecedor",
        "contrato": "Contrato",
        "id_nf": "Numero da NF",
        "item_medido": "Item medido",
        "valor_item": "Valor do item",
    }
    faltantes = [
        label
        for chave, label in campos_essenciais.items()
        if not valor_planilha_preenchido(registro.get(chave))
    ]
    if faltantes:
        raise ErroRegistroEsperado(
            "A linha possui campo(s) obrigatorio(s) vazio(s) ou invalido(s): "
            + ", ".join(faltantes)
            + ". Corrija a planilha e execute novamente."
        )


def ler_status(df, idx):
    for coluna in df.columns:
        chave = str(coluna).strip().lower()
        if chave == "status" or chave.startswith("status."):
            valor = df.loc[idx, coluna]
            if pd.notna(valor) and str(valor).strip():
                return str(valor)
    return ""


def atualizar_status_mensagem_no_excel(caminho_xlsx, updates):
    wb = load_workbook(caminho_xlsx)
    ws = wb.active
    col_status = garantir_coluna(ws, "status")
    col_mensagem = garantir_coluna(ws, "mensagem")
    for linha, status, mensagem in updates:
        ws.cell(row=int(linha), column=col_status).value = str(status)
        ws.cell(row=int(linha), column=col_mensagem).value = str(mensagem)
    wb.save(caminho_xlsx)


def atualizar_numero_solicitacao_no_excel(caminho_xlsx, updates):
    wb = load_workbook(caminho_xlsx)
    ws = wb.active
    col_solicitacao = garantir_coluna(ws, "numero_solicitacao")
    for linha, numero in updates:
        ws.cell(row=int(linha), column=col_solicitacao).value = str(numero)
    wb.save(caminho_xlsx)


def garantir_coluna(ws, nome):
    for coluna in range(1, ws.max_column + 1):
        valor = ws.cell(row=1, column=coluna).value
        if valor is not None and str(valor).strip().lower() == nome:
            return coluna
    coluna = ws.max_column + 1
    ws.cell(row=1, column=coluna).value = nome
    return coluna


def envio_final_habilitado(parametros_execucao):
    if "confirm_submission" in parametros_execucao:
        valor = parametros_execucao.get("confirm_submission")
        return str(valor).strip().lower() in {"1", "true", "t", "yes", "y", "on"}

    if "submit" in parametros_execucao:
        valor = parametros_execucao.get("submit")
        return str(valor).strip().lower() in {"1", "true", "t", "yes", "y", "on"}

    if "modo_producao" in parametros_execucao:
        valor = parametros_execucao.get("modo_producao")
        return str(valor).strip().lower() in {"1", "true", "t", "yes", "y", "on"}

    return True


def salvar_resumo(pasta_saida, planilha_trabalho, updates_excel, updates_solicitacao, resumo_execucao):
    resumo = {
        "planilha_saida": planilha_trabalho.name,
        "gerado_em": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "resumo_execucao": {
            "ok": resumo_execucao.get("ok", 0),
            "erro": resumo_execucao.get("erro", 0),
            "teste": resumo_execucao.get("teste", 0),
            "pulado_ok": resumo_execucao.get("pulado_ok", 0),
            "total_processado": resumo_execucao.get("total_processado", 0),
            "inicio_rotina": resumo_execucao.get("inicio_rotina", ""),
            "fim_rotina": resumo_execucao.get("fim_rotina", ""),
            "tempo_total_rotina_segundos": round(resumo_execucao.get("tempo_total_rotina_segundos", 0.0), 2),
            "tempo_total_rotina_formatado": formatar_duracao(resumo_execucao.get("tempo_total_rotina_segundos", 0.0)),
            "tempo_medio_registro_segundos": round(resumo_execucao.get("tempo_medio_registro_segundos", 0.0), 2),
            "tempo_medio_registro_formatado": formatar_duracao(resumo_execucao.get("tempo_medio_registro_segundos", 0.0)),
        },
        "status": updates_excel,
        "solicitacoes": updates_solicitacao,
    }
    salvar_resumo_em_aba_planilha(planilha_trabalho, resumo)


def salvar_resumo_em_aba_planilha(planilha_trabalho, resumo):
    wb = load_workbook(planilha_trabalho)
    remover_abas_credenciais(wb)
    nome_aba = "Resumo_Execucao"
    if nome_aba in wb.sheetnames:
        ws = wb[nome_aba]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(nome_aba)

    resumo_execucao = resumo.get("resumo_execucao", {})
    preencher_resumo_visual(ws, resumo_execucao, resumo.get("planilha_saida", ""))

    wb.save(planilha_trabalho)


def remover_abas_credenciais(workbook):
    nomes_alvo = {"credenciais", "login", "acesso", "usuario"}
    abas_para_remover = []

    for nome_aba in list(workbook.sheetnames):
        aba_normalizada = normalizar_chave_excel(nome_aba)
        if aba_normalizada in nomes_alvo:
            abas_para_remover.append(nome_aba)
            continue

        worksheet = workbook[nome_aba]
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            continue

        headers = {normalizar_chave_excel(valor) for valor in header_row if valor is not None}
        if {"usuario", "senha"}.issubset(headers) or {"username", "password"}.issubset(headers):
            abas_para_remover.append(nome_aba)

    for nome_aba in abas_para_remover:
        if nome_aba in workbook.sheetnames and len(workbook.sheetnames) > 1:
            del workbook[nome_aba]


def preencher_resumo_visual(ws, resumo_execucao, planilha_saida):
    cor_fundo = "F8FAFC"
    cor_titulo = "0F172A"
    cor_texto_claro = "FFFFFF"
    cor_borda = "CBD5E1"
    cor_label = "475569"
    cor_ok = "DCFCE7"
    cor_ok_texto = "166534"
    cor_erro = "FEE2E2"
    cor_erro_texto = "991B1B"
    cor_teste = "FEF3C7"
    cor_teste_texto = "92400E"
    cor_pulado = "DBEAFE"
    cor_pulado_texto = "1D4ED8"
    cor_neutra = "E2E8F0"
    cor_neutra_texto = "334155"

    thin = Side(border_style="thin", color=cor_borda)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 22

    for row in range(1, 20):
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill("solid", fgColor=cor_fundo)

    ws.merge_cells("A1:D1")
    titulo = ws["A1"]
    titulo.value = "Resumo da Execucao"
    titulo.fill = PatternFill("solid", fgColor=cor_titulo)
    titulo.font = Font(color=cor_texto_claro, bold=True, size=16)
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    titulo.border = border
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:D2")
    subtitulo = ws["A2"]
    subtitulo.value = f"Arquivo gerado: {planilha_saida}"
    subtitulo.fill = PatternFill("solid", fgColor="E2E8F0")
    subtitulo.font = Font(color="0F172A", italic=True, size=10)
    subtitulo.alignment = Alignment(horizontal="left", vertical="center")
    subtitulo.border = border

    cards = [
        ("A4", "OK", resumo_execucao.get("ok", 0), cor_ok, cor_ok_texto),
        ("B4", "ERRO", resumo_execucao.get("erro", 0), cor_erro, cor_erro_texto),
        ("C4", "TESTE", resumo_execucao.get("teste", 0), cor_teste, cor_teste_texto),
        ("D4", "PULADO OK", resumo_execucao.get("pulado_ok", 0), cor_pulado, cor_pulado_texto),
    ]

    for ref, label, valor, cor_bg, cor_texto in cards:
        cell = ws[ref]
        cell.value = f"{label}\n{valor}"
        cell.fill = PatternFill("solid", fgColor=cor_bg)
        cell.font = Font(color=cor_texto, bold=True, size=13)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.row_dimensions[cell.row].height = 40

    metricas = [
        (7, "Inicio da rotina", resumo_execucao.get("inicio_rotina", "")),
        (8, "Fim da rotina", resumo_execucao.get("fim_rotina", "")),
        (9, "Tempo total da rotina", resumo_execucao.get("tempo_total_rotina_formatado", "00:00:00")),
        (10, "Tempo medio por registro", resumo_execucao.get("tempo_medio_registro_formatado", "00:00:00")),
        (11, "Total processado", resumo_execucao.get("total_processado", 0)),
    ]

    for row, label, valor in metricas:
        label_cell = ws.cell(row=row, column=1)
        value_cell = ws.cell(row=row, column=2)
        label_cell.value = label
        value_cell.value = valor
        label_cell.fill = PatternFill("solid", fgColor=cor_neutra)
        value_cell.fill = PatternFill("solid", fgColor="FFFFFF")
        label_cell.font = Font(color=cor_label, bold=True)
        value_cell.font = Font(color=cor_neutra_texto, bold=True)
        label_cell.alignment = Alignment(horizontal="left", vertical="center")
        value_cell.alignment = Alignment(horizontal="left", vertical="center")
        label_cell.border = border
        value_cell.border = border

    ws.merge_cells("A13:D13")
    rodape = ws["A13"]
    rodape.value = "Os detalhes de erro e mensagem continuam disponiveis na aba principal da planilha."
    rodape.fill = PatternFill("solid", fgColor="E0F2FE")
    rodape.font = Font(color="0C4A6E", italic=True, size=10)
    rodape.alignment = Alignment(horizontal="left", vertical="center")
    rodape.border = border


def formatar_duracao(total_segundos):
    total_segundos = int(round(float(total_segundos or 0)))
    horas, resto = divmod(total_segundos, 3600)
    minutos, segundos = divmod(resto, 60)
    return f"{horas:02d}:{minutos:02d}:{segundos:02d}"


def coletar_amostra_modal(driver, modal_id, limite=5):
    if not modal_id:
        return []
    seletor = f"{modal_id} table tbody tr"
    return coletar_amostra_tabela(driver, seletor, limite=limite)


def coletar_amostra_tabela(driver, seletor, limite=5):
    amostra = []
    for linha in driver.find_elements(By.CSS_SELECTOR, seletor)[:limite]:
        try:
            texto = " | ".join(
                celula.text.strip()
                for celula in linha.find_elements(By.TAG_NAME, "td")
                if celula.text.strip()
            ).strip()
        except Exception:
            texto = ""
        if texto:
            amostra.append(texto)
    return amostra


def montar_erro_busca(label, termo_busca, sugestoes=None, detalhe_extra=""):
    mensagem = (
        f"{label} nao encontrado(a) no BlueEZ para o valor informado na planilha: '{termo_busca}'. "
        f"Revise o cadastro ou ajuste o texto enviado."
    )
    if detalhe_extra:
        mensagem += f" {detalhe_extra}"
    if sugestoes:
        mensagem += " Itens exibidos na busca: " + " || ".join(sugestoes)
    return mensagem


def obter_elemento_clicavel(driver, wait, seletor_css, descricao):
    try:
        return wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, seletor_css)))
    except TimeoutException as exc:
        candidatos = driver.find_elements(By.CSS_SELECTOR, seletor_css)
        detalhes = []
        if candidatos:
            for indice, elemento in enumerate(candidatos[:3], start=1):
                try:
                    detalhes.append(
                        (
                            f"candidato {indice}: exibido={elemento.is_displayed()} "
                            f"habilitado={elemento.is_enabled()} "
                            f"texto='{(elemento.text or '').strip()}' "
                            f"html='{(elemento.get_attribute('outerHTML') or '')[:180]}'"
                        )
                    )
                except Exception:
                    detalhes.append(f"candidato {indice}: estado indisponivel")
        else:
            detalhes.append("nenhum elemento encontrado com esse seletor")

        contexto = (
            f"url_atual={safe_current_url(driver)} | titulo={safe_title(driver)!r} | "
            f"seletor={seletor_css!r}"
        )
        raise ValueError(
            f"Nao foi possivel interagir com {descricao} no BlueEZ. "
            f"{contexto}. Detalhes: {' || '.join(detalhes)}"
        ) from exc


def safe_capability(driver, key):
    try:
        return driver.capabilities.get(key)
    except Exception:
        return None


def safe_current_url(driver):
    try:
        return driver.current_url
    except Exception:
        return "indisponivel"


def safe_title(driver):
    try:
        return driver.title
    except Exception:
        return "indisponivel"


def limpar_codigo_excel(valor):
    texto = "" if pd.isna(valor) else str(valor).strip()
    return texto[:-2] if texto.endswith(".0") else texto


def normaliza_nome(valor):
    return "".join(ch for ch in str(valor).lower() if ch.isalnum())


def normalizar_texto(valor):
    if valor is None:
        return ""
    texto = str(valor).replace("\n", " ").strip().lower()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(ch for ch in texto if unicodedata.category(ch) != "Mn")
    return re.sub(r"\s+", " ", texto)


def formatar_cnpj(valor):
    digitos = re.sub(r"\D+", "", str(valor or ""))
    if len(digitos) != 14:
        raise ValueError(f"CNPJ invalido: {valor!r}")
    return f"{digitos[:2]}.{digitos[2:5]}.{digitos[5:8]}/{digitos[8:12]}-{digitos[12:]}"


def formatar_data(valor):
    data = converter_data_excel(valor)
    if pd.isna(data):
        raise ValueError(f"Data invalida: {valor!r}")
    return data.strftime("%d/%m/%Y")


def formatar_competencia(valor):
    data = converter_data_excel(valor)
    if pd.isna(data):
        raise ValueError(f"Competencia invalida: {valor!r}")
    return data.strftime("%m/%Y")


def converter_data_excel(valor):
    if pd.isna(valor):
        return pd.NaT
    if isinstance(valor, (int, float)):
        numero = float(valor)
        if 1 <= numero <= 60000:
            return pd.to_datetime(numero, unit="D", origin="1899-12-30", errors="coerce")
    return pd.to_datetime(valor, dayfirst=True, errors="coerce")
